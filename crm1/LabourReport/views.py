from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages,auth
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from .models import *
from ProductivityReport.models import *
from .forms import *#SiteEngDayForm,UpdateForm,CreateUserForm,Area_Input,AddCont,Add_Labour,Add_Lab_To_Contractor,ResetPasswordForm
from ProductivityReport.forms import *
from datetime import date
from .decorators import allowed_users, unauthenticated_user
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font,Alignment
from datetime import datetime, timedelta
import pandas as pd

def Navbar(request):
    return render(request,'LabourReport/Navbar.html')

@unauthenticated_user
def LoginPage(request):
    if request.method == 'POST':
        userid = request.POST['userid']
        password = request.POST['pass']
        user_auth = authenticate(username=userid, password=password)
        admin = Group.objects.get(name="Admin").user_set.all()
        SE = Group.objects.get(name="Site Engineer").user_set.all()
        SLI = Group.objects.get(name="Site Labour Incharge").user_set.all()
        CLI = Group.objects.get(name="Camp Labour Incharge").user_set.all()
        Mang = Group.objects.get(name="Management").user_set.all()
        if user_auth is not None:
            if user_auth in SE:
                login(request,user_auth)
                return redirect('HomeSE')
            elif user_auth in admin:
                login(request,user_auth)
                return redirect('HomeAdmin')
            elif user_auth in SLI:
                login(request,user_auth)
                print("in")
                return redirect('HomeSLI')
            elif user_auth in CLI:
                login(request,user_auth)
                return redirect('HomeCLI')
            elif user_auth in Mang:
                login(request,user_auth)
                return redirect('HomeMang')
        else:
            messages.info(request,'Credential is incorrect')
            return redirect('/')
    return render(request,'LabourReport/LoginPage.html')

def LogoutUser(request):
    logout(request)
    return redirect('Login')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def HomeSE(request):
    return render(request,'LabourReport/HomeSE.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def AddDaySE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    # print("d",d1,d2)
    Report=SiteEngDay.objects.filter(created_at__range=[d1,d2],Areaname=Areaname_id).order_by('ContractorName')
    form=SiteEngDayForm()
    if request.method =='POST':
        form=SiteEngDayForm(request.POST)
        print(form.errors)
        if form.is_valid():
            form.save()
        return redirect('AddDaySE')
    return render(request,'LabourReport/SiteEngAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def ViewDaySE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2,type(d1))
    Report=SiteEngDay.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/SiteEngViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def ViewNightSE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SiteEngNight.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/SiteEngViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def AddNightSE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SiteEngNight.objects.filter(created_at__range=[d1,d2],Areaname=Areaname_id)
    form=SiteEngNightForm()
    if request.method =='POST':
        form=SiteEngNightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('/AddNightSE/')
    return render(request,'LabourReport/SiteEngAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def DeleteDaySE(request,i):
    new=SiteEngDay.objects.get(id=i)
    new.delete()
    return redirect('/AddDaySE/')
    

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def DeleteNightSE(request,i):
    new=SiteEngNight.objects.get(id=i)
    new.delete()
    return redirect('/AddNightSE/')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def DLRSummary(request):
    if request.method=='POST':
        # print(request.POST)
        query_dict = request.POST.dict()
        print(query_dict.keys())
        if 'Show' in query_dict.keys():
            # Get data from SiteEngDay table
            current_user = request.user
            Areaname = Area.objects.filter(Username=current_user.username)
            d1 = query_dict['From']
            d2 = query_dict['To']
            print(query_dict['Shift'])
            if query_dict['Shift']=='Day':
                data = SiteEngDay.objects.filter(created_at__range=[d1,d2],Areaname=Areaname[0].id).order_by('ContractorName')
            else:
                data = SiteEngNight.objects.filter(created_at__range=[d1,d2],Areaname=Areaname[0].id).order_by('ContractorName')
            df = pd.DataFrame(columns=['ContractorName','Date','LabourCategory','CategoryName','NoLabor'])
            for i in data:
                # print all data
                date = str(i.created_at)[:10]
                date = date[8:10]+"/"+date[5:7]+"/"+date[0:4]
                df.loc[len(df)] = {'ContractorName':i.ContractorName,'Date':date,'LabourCategory':i.LabourCategory,'CategoryName':i.CategoryName,'NoLabor':i.NoLabor}
                # df = df.append({'ContractorName':i.ContractorName,'Date':date,'LabourCategory':i.LabourCategory,'CategoryName':i.CategoryName,'NoLabor':i.NoLabor},ignore_index=True)
            # Get unique ContractorName from df
            cont_name = df['ContractorName'].unique()
            cont_row_span = {}
            for i in cont_name:
                cont_row_span[i] = len(df[df['ContractorName']==i])
            return render(request,'LabourReport/DLR_Summary.html',{'df':df,'from':d1,'to':d2,'Shift':query_dict['Shift'],'cont_row_span':cont_row_span})
        elif 'Export' in query_dict.keys():
            current_user = request.user
            Areaname = Area.objects.filter(Username=current_user.username)
            d1 = query_dict['From']
            d2 = query_dict['To']
            print(query_dict['Shift'])
            if query_dict['Shift']=='Day':
                data = SiteEngDay.objects.filter(created_at__range=[d1,d2],Areaname=Areaname[0].id).order_by('ContractorName')
            else:
                data = SiteEngNight.objects.filter(created_at__range=[d1,d2],Areaname=Areaname[0].id).order_by('ContractorName')
            df = pd.DataFrame(columns=['ContractorName','Date','LabourCategory','CategoryName','NoLabor'])
            for i in data:
                # print all data
                date = str(i.created_at)[:10]
                date = date[8:10]+"/"+date[5:7]+"/"+date[0:4]
                print(date)
                df.loc[len(df)] = {'ContractorName':i.ContractorName,'Date':date,'LabourCategory':i.LabourCategory,'CategoryName':i.CategoryName,'NoLabor':i.NoLabor}
                # df = df.append({'ContractorName':i.ContractorName,'Date':date,'LabourCategory':i.LabourCategory,'CategoryName':i.CategoryName,'NoLabor':i.NoLabor},ignore_index=True)
            # Create Workbook and add worksheet
            workbook = Workbook()
            # Get active worksheet/tab
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename="DLR "+d1+" to "+d2+" "+query_dict['Shift']+".xlsx"
            response['Content-Disposition'] = 'attachment; filename='+filename
            worksheet = workbook.active
            worksheet.title = 'DLR Summary'

            # Define the titles for columns
            columns = [
                'ContractorName',
                'Date',
                'LabourCategory',
                'CategoryName',
                'NoLabor',
            ]

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title
            
            # Bold the header
            for cell in worksheet["1:1"]:
                cell.font = Font(bold=True)
            row_num = 2
            # Append data in sheet
            for i in range(0,len(df)):
                # print(df.iloc[i])
                print(df.iloc[i]['ContractorName'])
                # Reverse the date format
                date = str(df.iloc[i]['Date'])
                row = [
                    str(df.iloc[i]['ContractorName']),
                    date,
                    str(df.iloc[i]['LabourCategory']),
                    str(df.iloc[i]['CategoryName']),
                    df.iloc[i]['NoLabor'],
                ]
                print(row)
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                row_num += 1
            # Save the file
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length+3
                worksheet.column_dimensions[column].width = adjusted_width
            workbook.save(response)
            return response

    return render(request,'LabourReport/DLR_Summary.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def HomeAdmin(request):
    # Get Data of users
    User_data=User.objects.all().order_by('username')
    User_data=User.objects.all().order_by('username')
    df = pd.DataFrame(columns=['username','email','group','id'])
    for i in User_data:
        if i.username != 'harshitkava':
            df.loc[len(df)] = {'username':i.username,'id':i.id,'email':i.email,'group':i.groups.all()[0]}

    Area_data=Area.objects.all().order_by('AreaName')
    # Convert to Dataframe
    df1=pd.DataFrame(Area_data.values())
    df1 = df1[['AreaName','Username']]
    # Merge Dataframes
    df2 = pd.merge(df,df1,how='left',left_on='username',right_on='Username')
    # drop Username column
    df2 = df2.drop(['Username'],axis=1)
    # print(df2)
    return render(request,'LabourReport/Admin/HomeAd.html',{'data':df2})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditUser(request,i):
    # Get Data of users
    i = float(i)
    User_data=User.objects.get(pk=i)
    Area_data=Area.objects.get(Username=User_data.username)
    registration_form=CreateUserForm(request.POST or None,instance=User_data)
    Area_input=Area_Input(request.POST or None,instance=Area_data)
    print("*************",registration_form.is_valid())
    if registration_form.is_valid() and Area_input.is_valid():
        registration_form.save()
        Area_input.save()
        print(1)
        return redirect('HomeAdmin')
    return render(request,'LabourReport/Admin/EditUser.html',{'form1':registration_form,'form2':Area_input})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteUser(request,i):
    i = float(i)
    User_data=User.objects.get(pk=i)
    User_data.delete()
    return redirect('HomeAdmin')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowContractor(request):
    data = ContractorDetail.objects.all().order_by('ContractorName')
    df = pd.DataFrame(columns=['ContractorName','ContractorNumber','id'])
    for i in data:
        df.loc[len(df)] = {'ContractorName':i.ContractorName,'ContractorNumber':i.ContractorNumber,'id':i.id}
        # df = df.append({'ContractorName':i.ContractorName,'ContractorNumber':i.ContractorNumber,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowContractor.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditContractor(request,i):
    i = float(i)
    data = ContractorDetail.objects.get(pk=i)
    form = AddCont(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('Contractor')
    return render(request,'LabourReport/Admin/EditContractor.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteContractor(request,i):
    i = float(i)
    data = ContractorDetail.objects.get(pk=i)
    data.delete()
    return redirect('Contractor')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowStructure(request):
    data = Structure.objects.all().order_by('StructureName')
    df = pd.DataFrame(columns=['StructureName','id'])
    for i in data:
        df.loc[len(df)] = {'StructureName':i.StructureName,'id':i.id}
        # df = df.append({'StructureName':i.StructureName,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowStructure.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditStructure(request,i):
    i = float(i)
    data = Structure.objects.get(pk=i)
    form = StructureForm(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowStructure')
    return render(request,'LabourReport/Admin/EditStructure.html',{'Form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteStructure(request,i):
    i = float(i)
    data = Structure.objects.get(pk=i)
    data.delete()
    return redirect('ShowStructure')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowLabours(request):
    data = AddLabour.objects.all().order_by('LabourCategory')
    df = pd.DataFrame(columns=['LabourCategory','id'])
    for i in data:
        df.loc[len(df)] = {'LabourCategory':i.LabourCategory,'id':i.id}
        # df = df.append({'LabourName':i.LabourName,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowLabours.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditLabours(request,i):
    i = float(i)
    data = AddLabour.objects.get(pk=i)
    form = Add_Labour(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowLabours')
    return render(request,'LabourReport/Admin/EditLabours.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteLabours(request,i):
    i = float(i)
    data = AddLabour.objects.get(pk=i)
    data.delete()
    return redirect('ShowLabours')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowLabourOfContractor(request):
    data = LabourOfContractor.objects.all()
    df = pd.DataFrame(columns=['ContractorName','LabourCategory','id'])
    for i in data:
        df.loc[len(df)] = {'ContractorName':str(i.ContractorName),'LabourCategory':i.LabourCategory,'id':i.id}
        # df = df.append({'ContractorName':i.ContractorName,'id':i.id},ignore_index=True)
    df = df.sort_values(by='ContractorName')
    # print(df.sort_values(by='ContractorName',ascending=True))
    return render(request,'LabourReport/Admin/ShowLabourOfContractor.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditLabourOfContractor(request,i):
    i = float(i)
    data = LabourOfContractor.objects.get(pk=i)
    form = Add_Lab_To_Contractor(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowLabourOfContractor')
    return render(request,'LabourReport/Admin/EditLabourOfContractor.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteLabourOfContractor(request,i):
    i = float(i)
    data = LabourOfContractor.objects.get(pk=i)
    data.delete()
    return redirect('ShowLabourOfContractor')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowActivity(request):
    data = CategoryOfDeployment.objects.all().order_by('ActivityName')
    df = pd.DataFrame(columns=['CategoryName','ActivityName','id'])
    for i in data:
        df.loc[len(df)] = {'CategoryName':i.CategoryName,'ActivityName':i.ActivityName,'id':i.id}
        # df = df.append({'ActivityName':i.ActivityName,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowActivity.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditActivity(request,i):
    i = float(i)
    data = CategoryOfDeployment.objects.get(pk=i)
    form = CategoryOfDeploymentForm(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowActivity')
    return render(request,'LabourReport/Admin/EditActivity.html',{'Form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def AddUser(request):
    registration_form=CreateUserForm()
    Area_input=Area_Input()
    # descending order of groups in User table
    User_data=User.objects.all().order_by('username')
    if request.method == 'POST':
        registration_form=CreateUserForm(request.POST)
        Area_input=Area_Input(request.POST)
        
        print(2)
        print(request.POST.get('groups'))
        print(registration_form.errors)
        if registration_form.is_valid():
            print(3)
            user=registration_form.save()
            if request.POST.get('groups')== "1":
                grp = Group.objects.get(name='Site Engineer')
                # grp=list(grp)
                user.groups.add(grp)
            elif request.POST.get('groups')== "2":
                grp = Group.objects.get(name='Site Labour Incharge')
                user.groups.add(grp)
            elif request.POST.get('groups')== "3":
                grp = Group.objects.get(name='Admin')
                user.groups.add(grp)
            elif request.POST.get('groups')== "4":
                grp = Group.objects.get(name='Management')
                user.groups.add(grp)
            elif request.POST.get('groups')== "5":
                grp = Group.objects.get(name='Camp Labour Incharge')
                user.groups.add(grp)
            if Area_input.is_valid():
                Area_input.save()
                print(4)
                return redirect('AddUser')
    return render(request,'LabourReport/Admin/User.html',{'form1':registration_form,'form2':Area_input,'User_data':User_data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def AddContractor(request):
    form=AddCont()
    data = ContractorDetail.objects.all()
    if request.method == 'POST':
        form=AddCont(request.POST)
        if form.is_valid():
            form.save()
        return render(request,'LabourReport/Admin/Contractor.html',{'form':form,'data':data})    
    return render(request,'LabourReport/Admin/Contractor.html',{'form':form,'data':data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])    
def AddLabours(request):
    form=Add_Labour()
    labour_data = AddLabour.objects.all()
    if request.method == 'POST':
        form=Add_Labour(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ShowLabours')
    
    return render(request,'LabourReport/Admin/Labour.html',{'form':form,'labour_data':labour_data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def LaboursOfContractor(request):
    form=Add_Lab_To_Contractor()
    data = LabourOfContractor.objects.all()
    if request.method == 'POST':
        form=Add_Lab_To_Contractor(request.POST)
        
        if form.is_valid():
            form.save()
            return render(request,'LabourReport/Admin/LabourToCont.html',{'form':form,'data':data})
    return render(request,'LabourReport/Admin/LabourToCont.html',{'form':form,'data':data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ResetPassword (request):
    form=ResetPasswordForm()
    if request.method == 'POST':
        form=ResetPasswordForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('HomeAdmin')
    return render(request,'LabourReport/Admin/ResetPassword.html',{'form':form})

@login_required(login_url='Login')
# @allowed_users(allowed_roles=['Site Engineer'])
def load_labour(request):
    contractor_id = request.GET.get('contractor_id')
    print(contractor_id)
    contractor_name = ContractorDetail.objects.get(id=contractor_id)
    labourofCont = LabourOfContractor.objects.filter(ContractorName=contractor_name).order_by('LabourCategory')
    labour =AddLabour.objects.filter().order_by('LabourCategory')
    print(labourofCont,labour)
    return render(request, 'LabourReport/Admin/labour_dropdown_list_options.html', {'labour': labour,'labourofCont':labourofCont})

@login_required(login_url='Login')
# @allowed_users(allowed_roles=['Site Engineer'])
def load_cat(request):
    Labour_id = request.GET.get('contractor_id')
    print(Labour_id)
    # CategoryOfDeployment_id = CategoryOfDeployment.objects.get(id=Labour_id)
    # print(CategoryOfDeployment_id)
    labour_name = LabourOfContractor.objects.get(id=Labour_id)
    print(labour_name,type(labour_name),str(labour_name))
    labour =AddLabour.objects.get(LabourCategory=labour_name)
    print(labour,type(labour))
    Category=CategoryOfDeployment.objects.filter(ActivityName=labour)
    print(Category,type(Category))
    # {'labour': labour,'labourofCont':labourofCont}
    return render(request, 'LabourReport/Admin/category_dropdown_list_options.html',{'Category':Category})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def HomeSLI(request):
    return render(request,'LabourReport/SLI/HomeSLI.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def AddDaySLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLIDay.objects.filter(created_at__range=[d1,d2])
    form=SLIDayForm()
    if request.method =='POST':
        form=SLIDayForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddDaySLI')
    return render(request,'LabourReport/SLI/SLIAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def DeleteDaySLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=SLIDay.objects.all()
    new=SLIDay.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=SLIDayForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddDaySLI')
    return render(request,'LabourReport/SLI/SLIDelDayData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def AddNightSLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLINight.objects.filter(created_at__range=[d1,d2])
    form=SLINightForm()
    if request.method =='POST':
        form=SLINightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddNightSLI')
    return render(request,'LabourReport/SLI/SLIAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def DeleteNightSLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=SLINight.objects.all()
    new=SLINight.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=SLINightForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddNightSLI')
    return render(request,'LabourReport/SLI/SLIDelNightData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def ViewDaySLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLIDay.objects.filter()
    return render(request,'LabourReport/SLI/SLIViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def ViewNightSLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLINight.objects.filter()
    return render(request,'LabourReport/SLI/SLIViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def ViewDayCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLIDay.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/CLI/CLIViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def ViewNightCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLINight.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/CLI/CLIViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def HomeCLI(request):
    return render(request,'LabourReport/CLI/HomeCLI.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def AddDayCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLIDay.objects.filter(created_at__range=[d1,d2])
    form=CLIDayForm()
    if request.method =='POST':
        form=CLIDayForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddDayCLI')
    return render(request,'LabourReport/CLI/CLIAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def DeleteDayCLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=CLIDay.objects.all()
    new=CLIDay.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=CLIDayForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddDayCLI')
    return render(request,'LabourReport/CLI/CLIDelDayData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def AddNightCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLINight.objects.filter(created_at__range=[d1,d2])
    form=CLINightForm()
    if request.method =='POST':
        form=CLINightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddNightCLI')
    return render(request,'LabourReport/CLI/CLIAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def DeleteNightCLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=CLINight.objects.all()
    new=CLINight.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=CLINightForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddNightCLI')
    return render(request,'LabourReport/CLI/CLIDelNightData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def HomeMang(request):
    return render(request,'LabourReport/Management/HomeMang.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def SiteReport(request):
    current_user = request.user
    if request.method == 'POST':
        shift = request.POST.get('shift')
        date = request.POST.get('date')
        date1=datetime.strptime(date, "%Y-%m-%d")
        tomorrow = date1 + timedelta(1)
        d1=date1.strftime("%Y-%m-%d")
        d2=tomorrow.strftime("%Y-%m-%d")
        print("d",d1,d2)
        # area = request.POST.get('area')
        # print(area)
        
        
        # print(shift,date,d1,d2,area,area_id)

        workbook = Workbook()
        # Get active worksheet/tab
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename="Deployment Report "+date+" "+shift+".xlsx"
        response['Content-Disposition'] = 'attachment; filename='+filename
        worksheet = workbook.active
        worksheet.title = 'Deployment Report'
        area_arr=["SBN","KV","DBM","RKP","MPZ","HBM","ALK","AIIMS","Casting Yard","Casting Yard QC","Casting Yard PM"]
        area_arr1=["Labour Category","","SBN","","KV","","DBM","","RKP","","MPZ","","HBM","","ALK","","AIIMS","","Casting Yard","","Casting Yard QC","","Casting Yard PM","","Total",""]
        area_arr2=["","","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI"]
        LabCat=[]

        # Deployment Sheets
        for col_num, column_title in enumerate(area_arr1[2:], 1):
                
            if column_title=="":
                    # merge with next cell
                worksheet.merge_cells(start_row=1, start_column=col_num-1+2, end_row=1, end_column=col_num+2)
                #     # align to center
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell = worksheet.cell(row=1, column=col_num+2)
                cell.value = column_title
            worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
            worksheet.cell(row=1, column=1).value = "Labour Category"
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        for col_num, column_title in enumerate(area_arr2[2:], 1):
            cell = worksheet.cell(row=2, column=col_num+2)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
        for cell in worksheet["2:2"]:
            cell.font = Font(bold=True)
        
        row_num = 3
        for i in area_arr:
            area=i
            area_id=Area.objects.filter(AreaName=area)
            id_list = []
            for i in area_id:
                id_list.append(i.id)
            if shift == 'Day':
                rows = SiteEngDay.objects.filter(Areaname__in=id_list,created_at__range=[d1,d2]).values_list('LabourCategory', 'NoLabor')
            elif shift == 'Night':
                rows = SiteEngNight.objects.filter(Areaname__in=id_list,created_at__range=[d1,d2]).values_list('LabourCategory', 'NoLabor')
            worksheet['A1'].font = Font(bold=True)
            # row_num = row_num+ 1
            for row in rows:
                
                LabourCategory = LabourOfContractor.objects.filter(pk=row[0]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                # Define the data for each cell in the row 
                row = [
                    LabourCategory[0],
                    row[1],
                ]
                # Assign the data for each cell of the row 
                # index of area in area_arr
                ind=area_arr1.index(area)+1
                if LabourCategory[0] in LabCat:
                    print(LabCat,area)
                    row_num1=LabCat.index(LabourCategory[0])+3
                    # print(int(worksheet.cell(row=row_num1, column=ind).value))
                    if worksheet.cell(row=row_num1, column=ind).value is None:
                        worksheet.cell(row=row_num1, column=ind).value = row[1]
                    else:
                        worksheet.cell(row=row_num1, column=ind).value = worksheet.cell(row=row_num1, column=ind).value + row[1]
                else:
                    print(2)
                    worksheet.cell(row=row_num, column=ind).value = row[1]
                    worksheet.cell(row=row_num, column=1).value = row[0]
                    worksheet.cell(row=row_num, column=1).font = Font(bold=True)
                    worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
                    row_num = row_num + 1
                    LabCat.append(LabourCategory[0])
                    # Total of row
                    worksheet.cell(row=row_num-1, column=25).value = '=SUM(C'+str(row_num-1)+'+E'+str(row_num-1)+'+G'+str(row_num-1)+'+I'+str(row_num-1)+'+K'+str(row_num-1)+'+M'+str(row_num-1)+'+O'+str(row_num-1)+'+Q'+str(row_num-1)+'+S'+str(row_num-1)+'+U'+str(row_num-1)+'+W'+str(row_num-1)+')'
                    worksheet.cell(row=row_num-1, column=26).value = '=SUM(D'+str(row_num-1)+'+F'+str(row_num-1)+'+H'+str(row_num-1)+'+J'+str(row_num-1)+'+L'+str(row_num-1)+'+N'+str(row_num-1)+'+P'+str(row_num-1)+'+R'+str(row_num-1)+'+T'+str(row_num-1)+'+V'+str(row_num-1)+'+X'+str(row_num-1)+')'
        # Total of column
        worksheet.cell(row=row_num, column=1).value = 'Total'
        worksheet.cell(row=row_num, column=1).font = Font(bold=True)
        worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        for i in range(3,27):
            # ascii of C
            worksheet.cell(row=row_num, column=i).value = '=SUM('+chr(64+i)+'3:'+chr(64+i)+str(row_num-1)+')'
        # freeze C3
        worksheet.freeze_panes = 'C3'
            
        
        
        overall_dict = {}

        # Site Sheets
        for i in area_arr:
            area=i
            area_id=Area.objects.filter(AreaName=area)
            id_list = []
            for i in area_id:
                id_list.append(i.id)
            # Create multiple sheets using openpyxl
            worksheet = workbook.create_sheet(area)
            columns = ['Site Name','Contractor Name', 'Labour Category', 'Category of Deployment','Structure', 'Deployment']
            row_num = 1
                # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
            # worksheet.title = area
            # worksheet = workbook.create_sheet(area)
            # xlfile = workbook.create_sheet(area)
            # count+=1
            # if count>1:
            #     xlfile = workbook.create_sheet(area)
            #     xlfile = workbook.active
            #     xlfile['A1'] = 'Date'

            # Define the titles for columns
            
            if shift == 'Day':
                rows = SiteEngDay.objects.filter(Areaname__in=id_list,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory','CategoryName','StructureName', 'NoLabor')
            elif shift == 'Night':
                rows = SiteEngNight.objects.filter(Areaname__in=id_list,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory','CategoryName','StructureName', 'NoLabor')
            
            worksheet['A1'].font = Font(bold=True)
            for cell in worksheet["1:1"]:
                cell.font = Font(bold=True)
            tot_lab=tot_help=tot_tot=0
            rows=list(rows)
            for row in rows:
                row_num += 1
                ContractorName = ContractorDetail.objects.filter(pk=row[0]).values_list('ContractorName')
                ContName=list(ContractorName[0])
                LabourCategory = LabourOfContractor.objects.filter(pk=row[1]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                CategoryName = CategoryOfDeployment.objects.filter(pk=row[2]).values_list('CategoryName')
                CategoryName =list(CategoryName[0])
                StructureName = Structure.objects.filter(pk=row[3]).values_list('StructureName')
                StructureName =list(StructureName[0])
                row=[
                    area,
                    ContName[0],
                    LabourCategory[0],
                    CategoryName[0],
                    StructureName[0],
                    row[4],
                ]
                print(row)
                if row[2] not in overall_dict:
                    overall_dict[row[2]] = row[5]
                else:
                    overall_dict[row[2]] += row[5]
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    print(cell_value)
                    cell.value = cell_value
            # write total formula in cell
            if len(rows)>0:
                worksheet['E'+str(row_num+1)] = 'Total'
                worksheet['F'+str(row_num+1)] = '=SUM(F2:F'+str(row_num)+')'
                worksheet['E'+str(row_num+1)].font = Font(bold=True)
            else:
                # merge cells
                worksheet.merge_cells('A'+str(row_num+1)+':E'+str(row_num+1))
                worksheet['A'+str(row_num+1)] = 'No Data Found'
                # align to center
                worksheet['A'+str(row_num+1)].alignment = Alignment(horizontal='center')
            
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length+3
                worksheet.column_dimensions[column].width = adjusted_width
        # print(row_num)
        # tabl = Table(displayName="Table1", ref="A1:F"+str(row_num))
        # style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
        #                showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        # tabl.tableStyleInfo = style
        # worksheet.add_table(tabl)

        worksheet = workbook.create_sheet('Overall')
        overall = SiteEngDay.objects.filter(created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory', 'NoLabor')
        print(1,overall)
        print(overall_dict)
        row_num = 1
        columns = ['Labour Category', 'Deployment']
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
        worksheet['A1'].font = Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
        for key,value in overall_dict.items():
            row_num += 1
            row=[
                key,
                value,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length+3
                worksheet.column_dimensions[column].width = adjusted_width
        workbook.save(response)
        return response

    return render(request,'LabourReport/Management/SiteReport.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def FinalReport(request):
    if request.method == 'POST':
        shift = request.POST.get('shift')
        date = request.POST.get('date')
        date1=datetime.strptime(date, "%Y-%m-%d")
        tomorrow = date1 + timedelta(1)
        d1=date1.strftime("%Y-%m-%d")
        d2=tomorrow.strftime("%Y-%m-%d")
        reporttype = request.POST.get('type')
        
        print(shift,date,d1,d2,reporttype)

        workbook = Workbook()
        
        
        if reporttype == 'Site':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename="Site Report "+date+" "+shift+".xlsx"
            response['Content-Disposition'] = 'attachment; filename='+filename
            worksheet = workbook.active
            worksheet.title = 'Site Report'+date+' '+shift
            area=["SBN","KV","DBM","RKP","MPZ","HBM","ALK","AIIMS","Casting Yard","Casting Yard QC","Casting Yard PM"]

            columns = ['Contractor Name', 'Labour Category', 'Category of Deployment','Structure', 'Deployment']
            worksheet.column_dimensions['A'].width = len(columns[0])
            worksheet.column_dimensions['B'].width = len(columns[1])

            worksheet['A1'] = 'Contractor Name'
            worksheet['B1'] = 'Types of Labour'
            cureent_cell=worksheet.cell(row=1,column=ord('A')-64)
            cureent_cell.alignment = Alignment(horizontal='center', vertical='center')
            cureent_cell=worksheet.cell(row=1,column=ord('B')-64)
            cureent_cell.alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(response)
        # return response
    return render(request,'LabourReport/Management/FinalReport.html')